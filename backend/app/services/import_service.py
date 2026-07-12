"""
Import Service for bulk curriculum data imports from Excel files

Provides methods for:
- Excel file parsing using openpyxl
- Import data validation (structural and semantic)
- Preview generation of entities to be created
- Transactional import execution with rollback capability
- Integration with ValidationService and AuditService

Requirements: 9.1-9.10
"""
from typing import Dict, Any, List, Optional, Tuple
from datetime import datetime, timezone
from sqlalchemy.orm import Session
from sqlalchemy.exc import SQLAlchemyError
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.models.university import University
from app.models.campus import Campus
from app.models.study_program import StudyProgram, university_programs
from app.models.academic_track import AcademicTrack, TrackLevel
from app.models.semester import Semester
from app.models.teaching_unit import TeachingUnit
from app.models.course import Course, course_prerequisites
from app.services.validation_service import ValidationService
from app.services.audit_service import AuditService


class ImportError:
    """Represents an import validation error with row number"""
    
    def __init__(self, row: Optional[int], sheet: str, message: str, error_type: str = "validation"):
        self.row = row
        self.sheet = sheet
        self.message = message
        self.error_type = error_type
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            "row": self.row,
            "sheet": self.sheet,
            "message": self.message,
            "type": self.error_type
        }


class ImportService:
    """Service for bulk curriculum import from Excel files"""
    
    def __init__(self, db: Session):
        self.db = db
        self.validation_service = ValidationService(db)
        self.audit_service = AuditService(db)
        self.errors: List[ImportError] = []
    
    def parse_excel_file(self, file_path: str) -> Dict[str, Any]:
        """
        Parse Excel file into structured data.
        
        Requirements: 9.1, 9.9
        
        Expected Excel structure:
        - Sheet "Universities": name, name_de, country, description
        - Sheet "Campuses": university_name, name, name_de, location, description
        - Sheet "Programs": name, name_de, code, description
        - Sheet "University_Programs": university_name, program_name
        - Sheet "Tracks": program_name, name, name_de, level, total_ects, description
        - Sheet "Semesters": track_name, name, semester_number, ects_required, description
        - Sheet "TeachingUnits": semester_name, name, name_de, code, ects_required, description
        - Sheet "Courses": semester_name, teaching_unit_name, name, name_de, code, ects_credits, coefficient, difficulty_level, description
        - Sheet "Prerequisites": course_name, prerequisite_name
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Dictionary with parsed data organized by entity type
            
        Raises:
            ValueError: If file cannot be opened or has invalid structure
        """
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            raise ValueError(f"Failed to open Excel file: {str(e)}")
        
        import_data = {
            "universities": [],
            "campuses": [],
            "programs": [],
            "university_programs": [],
            "tracks": [],
            "semesters": [],
            "teaching_units": [],
            "courses": [],
            "prerequisites": []
        }
        
        # Parse Universities
        if "Universities" in workbook.sheetnames:
            import_data["universities"] = self._parse_universities_sheet(workbook["Universities"])
        
        # Parse Campuses
        if "Campuses" in workbook.sheetnames:
            import_data["campuses"] = self._parse_campuses_sheet(workbook["Campuses"])
        
        # Parse Programs
        if "Programs" in workbook.sheetnames:
            import_data["programs"] = self._parse_programs_sheet(workbook["Programs"])
        
        # Parse University-Program links
        if "University_Programs" in workbook.sheetnames:
            import_data["university_programs"] = self._parse_university_programs_sheet(
                workbook["University_Programs"]
            )
        
        # Parse Academic Tracks
        if "Tracks" in workbook.sheetnames:
            import_data["tracks"] = self._parse_tracks_sheet(workbook["Tracks"])
        
        # Parse Semesters
        if "Semesters" in workbook.sheetnames:
            import_data["semesters"] = self._parse_semesters_sheet(workbook["Semesters"])
        
        # Parse Teaching Units
        if "TeachingUnits" in workbook.sheetnames:
            import_data["teaching_units"] = self._parse_teaching_units_sheet(
                workbook["TeachingUnits"]
            )
        
        # Parse Courses
        if "Courses" in workbook.sheetnames:
            import_data["courses"] = self._parse_courses_sheet(workbook["Courses"])
        
        # Parse Prerequisites
        if "Prerequisites" in workbook.sheetnames:
            import_data["prerequisites"] = self._parse_prerequisites_sheet(
                workbook["Prerequisites"]
            )
        
        workbook.close()
        
        return import_data
    
    def _parse_universities_sheet(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """Parse Universities sheet"""
        universities = []
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        
        for i, row in enumerate(rows, start=2):
            if not row or not any(row):  # Skip empty rows
                continue
            
            universities.append({
                "name": self._get_cell_value(row, 0),
                "name_de": self._get_cell_value(row, 1),
                "country": self._get_cell_value(row, 2) or "Germany",
                "description": self._get_cell_value(row, 3),
                "description_de": self._get_cell_value(row, 4),
                "_row": i
            })
        
        return universities
    
    def _parse_campuses_sheet(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """Parse Campuses sheet"""
        campuses = []
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        
        for i, row in enumerate(rows, start=2):
            if not row or not any(row):
                continue
            
            campuses.append({
                "university_name": self._get_cell_value(row, 0),
                "name": self._get_cell_value(row, 1),
                "name_de": self._get_cell_value(row, 2),
                "location": self._get_cell_value(row, 3),
                "description": self._get_cell_value(row, 4),
                "description_de": self._get_cell_value(row, 5),
                "_row": i
            })
        
        return campuses
    
    def _parse_programs_sheet(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """Parse Programs sheet"""
        programs = []
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        
        for i, row in enumerate(rows, start=2):
            if not row or not any(row):
                continue
            
            programs.append({
                "name": self._get_cell_value(row, 0),
                "name_de": self._get_cell_value(row, 1),
                "code": self._get_cell_value(row, 2),
                "description": self._get_cell_value(row, 3),
                "description_de": self._get_cell_value(row, 4),
                "_row": i
            })
        
        return programs
    
    def _parse_university_programs_sheet(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """Parse University_Programs sheet"""
        links = []
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        
        for i, row in enumerate(rows, start=2):
            if not row or not any(row):
                continue
            
            links.append({
                "university_name": self._get_cell_value(row, 0),
                "program_name": self._get_cell_value(row, 1),
                "_row": i
            })
        
        return links
    
    def _parse_tracks_sheet(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """Parse Tracks sheet"""
        tracks = []
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        
        for i, row in enumerate(rows, start=2):
            if not row or not any(row):
                continue
            
            tracks.append({
                "program_name": self._get_cell_value(row, 0),
                "name": self._get_cell_value(row, 1),
                "name_de": self._get_cell_value(row, 2),
                "level": self._get_cell_value(row, 3),
                "total_ects_required": self._get_cell_value(row, 4, int),
                "description": self._get_cell_value(row, 5),
                "description_de": self._get_cell_value(row, 6),
                "graduation_conditions": self._get_cell_value(row, 7),
                "_row": i
            })
        
        return tracks
    
    def _parse_semesters_sheet(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """Parse Semesters sheet"""
        semesters = []
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        
        for i, row in enumerate(rows, start=2):
            if not row or not any(row):
                continue
            
            semesters.append({
                "track_name": self._get_cell_value(row, 0),
                "name": self._get_cell_value(row, 1),
                "name_de": self._get_cell_value(row, 2),
                "semester_number": self._get_cell_value(row, 3, int),
                "ects_required": self._get_cell_value(row, 4, int),
                "description": self._get_cell_value(row, 5),
                "description_de": self._get_cell_value(row, 6),
                "_row": i
            })
        
        return semesters
    
    def _parse_teaching_units_sheet(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """Parse TeachingUnits sheet"""
        teaching_units = []
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        
        for i, row in enumerate(rows, start=2):
            if not row or not any(row):
                continue
            
            teaching_units.append({
                "semester_name": self._get_cell_value(row, 0),
                "name": self._get_cell_value(row, 1),
                "name_de": self._get_cell_value(row, 2),
                "code": self._get_cell_value(row, 3),
                "ects_required": self._get_cell_value(row, 4, int),
                "description": self._get_cell_value(row, 5),
                "description_de": self._get_cell_value(row, 6),
                "_row": i
            })
        
        return teaching_units
    
    def _parse_courses_sheet(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """Parse Courses sheet"""
        courses = []
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        
        for i, row in enumerate(rows, start=2):
            if not row or not any(row):
                continue
            
            courses.append({
                "semester_name": self._get_cell_value(row, 0),
                "teaching_unit_name": self._get_cell_value(row, 1),
                "name": self._get_cell_value(row, 2),
                "name_de": self._get_cell_value(row, 3),
                "code": self._get_cell_value(row, 4),
                "ects_credits": self._get_cell_value(row, 5, int),
                "coefficient": self._get_cell_value(row, 6, float),
                "difficulty_level": self._get_cell_value(row, 7, int),
                "description": self._get_cell_value(row, 8),
                "description_de": self._get_cell_value(row, 9),
                "_row": i
            })
        
        return courses
    
    def _parse_prerequisites_sheet(self, sheet: Worksheet) -> List[Dict[str, Any]]:
        """Parse Prerequisites sheet"""
        prerequisites = []
        rows = list(sheet.iter_rows(min_row=2, values_only=True))
        
        for i, row in enumerate(rows, start=2):
            if not row or not any(row):
                continue
            
            prerequisites.append({
                "course_name": self._get_cell_value(row, 0),
                "prerequisite_name": self._get_cell_value(row, 1),
                "_row": i
            })
        
        return prerequisites
    
    def _get_cell_value(self, row: tuple, index: int, convert_type=None):
        """
        Get cell value with optional type conversion.
        
        Args:
            row: Row tuple from openpyxl
            index: Column index
            convert_type: Optional type conversion function (int, float, str)
            
        Returns:
            Cell value, converted if specified, or None
        """
        if index >= len(row):
            return None
        
        value = row[index]
        
        if value is None or (isinstance(value, str) and value.strip() == ""):
            return None
        
        if convert_type:
            try:
                return convert_type(value)
            except (ValueError, TypeError):
                return None
        
        return value if not isinstance(value, str) else value.strip()
    
    async def validate_import_data(self, import_data: Dict[str, Any]) -> Tuple[bool, List[Dict[str, Any]]]:
        """
        Validate import data for structural and semantic correctness.
        
        Requirements: 9.2, 9.3, 9.5, 9.6
        
        Args:
            import_data: Dictionary with parsed import data
            
        Returns:
            Tuple of (is_valid, list of error dictionaries with row numbers and messages)
        """
        self.errors = []
        
        # Structural validation
        self._validate_universities_structure(import_data.get("universities", []))
        self._validate_programs_structure(import_data.get("programs", []))
        self._validate_tracks_structure(import_data.get("tracks", []))
        self._validate_semesters_structure(import_data.get("semesters", []))
        self._validate_courses_structure(import_data.get("courses", []))
        
        # Check for duplicates against existing data
        await self._validate_no_duplicate_universities(import_data)
        await self._validate_no_duplicate_programs(import_data)
        await self._validate_no_duplicate_campuses(import_data)
        await self._validate_no_duplicate_tracks(import_data)
        
        # Semantic validation (references, business rules)
        await self._validate_campuses_references(import_data)
        await self._validate_university_program_references(import_data)
        await self._validate_track_references(import_data)
        await self._validate_semester_references(import_data)
        await self._validate_teaching_unit_references(import_data)
        await self._validate_course_references(import_data)
        await self._validate_prerequisite_references(import_data)
        
        # Business rule validation
        await self._validate_prerequisite_chains(import_data)
        await self._validate_ects_consistency(import_data)
        
        is_valid = len(self.errors) == 0
        error_dicts = [e.to_dict() for e in self.errors]
        
        return is_valid, error_dicts
    
    async def _validate_no_duplicate_universities(self, import_data: Dict[str, Any]):
        """Check if universities already exist in database"""
        for uni in import_data.get("universities", []):
            uni_name = uni.get("name")
            if not uni_name:
                continue
            
            existing = self.db.query(University).filter(
                University.name == uni_name,
                University.is_deleted == False
            ).first()
            
            if existing:
                self.errors.append(ImportError(
                    uni.get("_row"),
                    "Universities",
                    f"University '{uni_name}' already exists in database. Please remove it from import file or update existing one.",
                    "duplicate"
                ))
    
    async def _validate_no_duplicate_programs(self, import_data: Dict[str, Any]):
        """Check if programs already exist in database"""
        for prog in import_data.get("programs", []):
            prog_name = prog.get("name")
            if not prog_name:
                continue
            
            existing = self.db.query(StudyProgram).filter(
                StudyProgram.name == prog_name,
                StudyProgram.is_deleted == False
            ).first()
            
            if existing:
                self.errors.append(ImportError(
                    prog.get("_row"),
                    "Programs",
                    f"Program '{prog_name}' already exists in database. Please remove it from import file or update existing one.",
                    "duplicate"
                ))
    
    async def _validate_no_duplicate_campuses(self, import_data: Dict[str, Any]):
        """Check if campuses already exist in database"""
        for campus in import_data.get("campuses", []):
            campus_name = campus.get("name")
            uni_name = campus.get("university_name")
            if not campus_name:
                continue
            
            # Check by name and university
            existing = self.db.query(Campus).join(University).filter(
                Campus.name == campus_name,
                University.name == uni_name,
                Campus.is_deleted == False,
                University.is_deleted == False
            ).first()
            
            if existing:
                self.errors.append(ImportError(
                    campus.get("_row"),
                    "Campuses",
                    f"Campus '{campus_name}' for university '{uni_name}' already exists in database.",
                    "duplicate"
                ))
    
    async def _validate_no_duplicate_tracks(self, import_data: Dict[str, Any]):
        """Check if tracks already exist in database"""
        for track in import_data.get("tracks", []):
            track_name = track.get("name")
            prog_name = track.get("program_name")
            if not track_name:
                continue
            
            # Check by name and program
            existing = self.db.query(AcademicTrack).join(StudyProgram).filter(
                AcademicTrack.name == track_name,
                StudyProgram.name == prog_name,
                AcademicTrack.is_deleted == False,
                StudyProgram.is_deleted == False
            ).first()
            
            if existing:
                self.errors.append(ImportError(
                    track.get("_row"),
                    "Tracks",
                    f"Track '{track_name}' for program '{prog_name}' already exists in database.",
                    "duplicate"
                ))
    
    def _validate_universities_structure(self, universities: List[Dict[str, Any]]):
        """Validate universities have required fields"""
        for uni in universities:
            if not uni.get("name"):
                self.errors.append(ImportError(
                    uni.get("_row"),
                    "Universities",
                    "University name is required"
                ))
    
    def _validate_programs_structure(self, programs: List[Dict[str, Any]]):
        """Validate programs have required fields"""
        for prog in programs:
            if not prog.get("name"):
                self.errors.append(ImportError(
                    prog.get("_row"),
                    "Programs",
                    "Program name is required"
                ))
    
    def _validate_tracks_structure(self, tracks: List[Dict[str, Any]]):
        """Validate tracks have required fields"""
        valid_levels = ["bachelor", "master", "doctorate"]
        
        for track in tracks:
            if not track.get("name"):
                self.errors.append(ImportError(
                    track.get("_row"),
                    "Tracks",
                    "Track name is required"
                ))
            
            level = track.get("level")
            if level and level.lower() not in valid_levels:
                self.errors.append(ImportError(
                    track.get("_row"),
                    "Tracks",
                    f"Invalid track level '{level}'. Must be one of: {', '.join(valid_levels)}"
                ))
            
            ects = track.get("total_ects_required")
            if ects is not None and (not isinstance(ects, int) or ects < 1):
                self.errors.append(ImportError(
                    track.get("_row"),
                    "Tracks",
                    "Total ECTS required must be a positive integer"
                ))
    
    def _validate_semesters_structure(self, semesters: List[Dict[str, Any]]):
        """Validate semesters have required fields"""
        for sem in semesters:
            if not sem.get("name"):
                self.errors.append(ImportError(
                    sem.get("_row"),
                    "Semesters",
                    "Semester name is required"
                ))
            
            sem_num = sem.get("semester_number")
            if sem_num is not None and (not isinstance(sem_num, int) or sem_num < 1):
                self.errors.append(ImportError(
                    sem.get("_row"),
                    "Semesters",
                    "Semester number must be a positive integer"
                ))
    
    def _validate_courses_structure(self, courses: List[Dict[str, Any]]):
        """Validate courses have required fields and valid values"""
        for course in courses:
            row = course.get("_row")
            
            if not course.get("name"):
                self.errors.append(ImportError(
                    row, "Courses", "Course name is required"
                ))
            
            # Validate ECTS (1-30)
            ects = course.get("ects_credits")
            if ects is None:
                self.errors.append(ImportError(
                    row, "Courses", "ECTS credits are required"
                ))
            elif not isinstance(ects, int) or ects < 1 or ects > 30:
                self.errors.append(ImportError(
                    row, "Courses", "ECTS credits must be an integer between 1 and 30"
                ))
            
            # Validate coefficient (0.1-10.0)
            coef = course.get("coefficient")
            if coef is None:
                self.errors.append(ImportError(
                    row, "Courses", "Coefficient is required"
                ))
            elif not isinstance(coef, (int, float)) or coef < 0.1 or coef > 10.0:
                self.errors.append(ImportError(
                    row, "Courses", "Coefficient must be a number between 0.1 and 10.0"
                ))
            
            # Validate difficulty (1-5)
            diff = course.get("difficulty_level")
            if diff is None:
                self.errors.append(ImportError(
                    row, "Courses", "Difficulty level is required"
                ))
            elif not isinstance(diff, int) or diff < 1 or diff > 5:
                self.errors.append(ImportError(
                    row, "Courses", "Difficulty level must be an integer between 1 and 5"
                ))
    
    async def _validate_campuses_references(self, import_data: Dict[str, Any]):
        """Validate campus university references"""
        university_names = {u.get("name") for u in import_data.get("universities", []) if u.get("name")}
        
        # Also check existing universities in DB
        existing_unis = self.db.query(University.name).filter(
            University.is_deleted == False
        ).all()
        university_names.update([u[0] for u in existing_unis])
        
        for campus in import_data.get("campuses", []):
            uni_name = campus.get("university_name")
            if uni_name and uni_name not in university_names:
                self.errors.append(ImportError(
                    campus.get("_row"),
                    "Campuses",
                    f"Referenced university '{uni_name}' not found"
                ))
    
    async def _validate_university_program_references(self, import_data: Dict[str, Any]):
        """Validate university-program link references"""
        university_names = {u.get("name") for u in import_data.get("universities", []) if u.get("name")}
        program_names = {p.get("name") for p in import_data.get("programs", []) if p.get("name")}
        
        # Also check existing entities
        existing_unis = self.db.query(University.name).filter(
            University.is_deleted == False
        ).all()
        university_names.update([u[0] for u in existing_unis])
        
        existing_progs = self.db.query(StudyProgram.name).filter(
            StudyProgram.is_deleted == False
        ).all()
        program_names.update([p[0] for p in existing_progs])
        
        for link in import_data.get("university_programs", []):
            uni_name = link.get("university_name")
            prog_name = link.get("program_name")
            
            if uni_name and uni_name not in university_names:
                self.errors.append(ImportError(
                    link.get("_row"),
                    "University_Programs",
                    f"Referenced university '{uni_name}' not found"
                ))
            
            if prog_name and prog_name not in program_names:
                self.errors.append(ImportError(
                    link.get("_row"),
                    "University_Programs",
                    f"Referenced program '{prog_name}' not found"
                ))
    
    async def _validate_track_references(self, import_data: Dict[str, Any]):
        """Validate track program references"""
        program_names = {p.get("name") for p in import_data.get("programs", []) if p.get("name")}
        
        existing_progs = self.db.query(StudyProgram.name).filter(
            StudyProgram.is_deleted == False
        ).all()
        program_names.update([p[0] for p in existing_progs])
        
        for track in import_data.get("tracks", []):
            prog_name = track.get("program_name")
            if prog_name and prog_name not in program_names:
                self.errors.append(ImportError(
                    track.get("_row"),
                    "Tracks",
                    f"Referenced program '{prog_name}' not found"
                ))
    
    async def _validate_semester_references(self, import_data: Dict[str, Any]):
        """Validate semester track references"""
        track_names = {t.get("name") for t in import_data.get("tracks", []) if t.get("name")}
        
        existing_tracks = self.db.query(AcademicTrack.name).filter(
            AcademicTrack.is_deleted == False
        ).all()
        track_names.update([t[0] for t in existing_tracks])
        
        for semester in import_data.get("semesters", []):
            track_name = semester.get("track_name")
            if track_name and track_name not in track_names:
                self.errors.append(ImportError(
                    semester.get("_row"),
                    "Semesters",
                    f"Referenced track '{track_name}' not found"
                ))
    
    async def _validate_teaching_unit_references(self, import_data: Dict[str, Any]):
        """Validate teaching unit semester references"""
        semester_names = {s.get("name") for s in import_data.get("semesters", []) if s.get("name")}
        
        existing_semesters = self.db.query(Semester.name).filter(
            Semester.is_deleted == False
        ).all()
        semester_names.update([s[0] for s in existing_semesters])
        
        for tu in import_data.get("teaching_units", []):
            sem_name = tu.get("semester_name")
            if sem_name and sem_name not in semester_names:
                self.errors.append(ImportError(
                    tu.get("_row"),
                    "TeachingUnits",
                    f"Referenced semester '{sem_name}' not found"
                ))
    
    async def _validate_course_references(self, import_data: Dict[str, Any]):
        """Validate course semester and teaching unit references"""
        semester_names = {s.get("name") for s in import_data.get("semesters", []) if s.get("name")}
        tu_names = {tu.get("name") for tu in import_data.get("teaching_units", []) if tu.get("name")}
        
        existing_semesters = self.db.query(Semester.name).filter(
            Semester.is_deleted == False
        ).all()
        semester_names.update([s[0] for s in existing_semesters])
        
        existing_tus = self.db.query(TeachingUnit.name).filter(
            TeachingUnit.is_deleted == False
        ).all()
        tu_names.update([tu[0] for tu in existing_tus])
        
        for course in import_data.get("courses", []):
            sem_name = course.get("semester_name")
            if sem_name and sem_name not in semester_names:
                self.errors.append(ImportError(
                    course.get("_row"),
                    "Courses",
                    f"Referenced semester '{sem_name}' not found"
                ))
            
            tu_name = course.get("teaching_unit_name")
            if tu_name and tu_name not in tu_names:
                self.errors.append(ImportError(
                    course.get("_row"),
                    "Courses",
                    f"Referenced teaching unit '{tu_name}' not found"
                ))
    
    async def _validate_prerequisite_references(self, import_data: Dict[str, Any]):
        """Validate prerequisite course references"""
        course_names = {c.get("name") for c in import_data.get("courses", []) if c.get("name")}
        
        existing_courses = self.db.query(Course.name).filter(
            Course.is_deleted == False
        ).all()
        course_names.update([c[0] for c in existing_courses])
        
        for prereq in import_data.get("prerequisites", []):
            course_name = prereq.get("course_name")
            prereq_name = prereq.get("prerequisite_name")
            
            if course_name and course_name not in course_names:
                self.errors.append(ImportError(
                    prereq.get("_row"),
                    "Prerequisites",
                    f"Referenced course '{course_name}' not found"
                ))
            
            if prereq_name and prereq_name not in course_names:
                self.errors.append(ImportError(
                    prereq.get("_row"),
                    "Prerequisites",
                    f"Referenced prerequisite course '{prereq_name}' not found"
                ))
    
    async def _validate_prerequisite_chains(self, import_data: Dict[str, Any]):
        """
        Validate prerequisite chains don't contain circular dependencies.
        
        Requirements: 9.3, 7.2
        """
        # Build course name to data map
        courses_map = {c.get("name"): c for c in import_data.get("courses", []) if c.get("name")}
        
        # Build prerequisite relationships
        prereq_graph = {}
        for prereq in import_data.get("prerequisites", []):
            course_name = prereq.get("course_name")
            prereq_name = prereq.get("prerequisite_name")
            
            if course_name and prereq_name:
                if course_name not in prereq_graph:
                    prereq_graph[course_name] = []
                prereq_graph[course_name].append(prereq_name)
        
        # Check for cycles using DFS
        def has_cycle(node, visited, rec_stack, path):
            visited.add(node)
            rec_stack.add(node)
            path.append(node)
            
            for neighbor in prereq_graph.get(node, []):
                if neighbor not in visited:
                    if has_cycle(neighbor, visited, rec_stack, path):
                        return True
                elif neighbor in rec_stack:
                    # Cycle detected
                    cycle_start = path.index(neighbor)
                    cycle_path = " -> ".join(path[cycle_start:] + [neighbor])
                    self.errors.append(ImportError(
                        None,
                        "Prerequisites",
                        f"Circular dependency detected: {cycle_path}",
                        "circular_dependency"
                    ))
                    return True
            
            path.pop()
            rec_stack.remove(node)
            return False
        
        visited = set()
        for course_name in prereq_graph.keys():
            if course_name not in visited:
                has_cycle(course_name, visited, set(), [])
    
    async def _validate_ects_consistency(self, import_data: Dict[str, Any]):
        """
        Validate ECTS consistency across tracks and courses.
        
        Requirements: 9.3
        """
        # Group courses by track to validate ECTS totals
        tracks = import_data.get("tracks", [])
        semesters = import_data.get("semesters", [])
        courses = import_data.get("courses", [])
        
        # Build mappings
        track_to_semesters = {}
        semester_to_courses = {}
        
        for semester in semesters:
            track_name = semester.get("track_name")
            if track_name:
                if track_name not in track_to_semesters:
                    track_to_semesters[track_name] = []
                track_to_semesters[track_name].append(semester.get("name"))
        
        for course in courses:
            semester_name = course.get("semester_name")
            if semester_name:
                if semester_name not in semester_to_courses:
                    semester_to_courses[semester_name] = []
                semester_to_courses[semester_name].append(course)
        
        # Validate each track
        for track in tracks:
            track_name = track.get("name")
            total_ects_required = track.get("total_ects_required")
            
            if not track_name or not total_ects_required:
                continue
            
            # Calculate total ECTS from courses
            track_semesters = track_to_semesters.get(track_name, [])
            total_ects = 0
            
            for sem_name in track_semesters:
                sem_courses = semester_to_courses.get(sem_name, [])
                for course in sem_courses:
                    ects = course.get("ects_credits")
                    if ects:
                        total_ects += ects
            
            # Allow some tolerance (courses might be optional)
            if total_ects > 0 and total_ects < total_ects_required * 0.8:
                self.errors.append(ImportError(
                    track.get("_row"),
                    "Tracks",
                    f"Track '{track_name}' requires {total_ects_required} ECTS but courses total only {total_ects} ECTS",
                    "ects_mismatch"
                ))
    
    async def preview_import(self, import_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Generate preview of entities to be created.
        
        Requirements: 9.4
        
        Args:
            import_data: Dictionary with validated import data
            
        Returns:
            Dictionary with counts and sample entities to be created
        """
        preview = {
            "universities": {
                "count": len(import_data.get("universities", [])),
                "samples": import_data.get("universities", [])[:3]
            },
            "campuses": {
                "count": len(import_data.get("campuses", [])),
                "samples": import_data.get("campuses", [])[:3]
            },
            "programs": {
                "count": len(import_data.get("programs", [])),
                "samples": import_data.get("programs", [])[:3]
            },
            "university_programs": {
                "count": len(import_data.get("university_programs", [])),
                "samples": import_data.get("university_programs", [])[:3]
            },
            "tracks": {
                "count": len(import_data.get("tracks", [])),
                "samples": import_data.get("tracks", [])[:3]
            },
            "semesters": {
                "count": len(import_data.get("semesters", [])),
                "samples": import_data.get("semesters", [])[:3]
            },
            "teaching_units": {
                "count": len(import_data.get("teaching_units", [])),
                "samples": import_data.get("teaching_units", [])[:3]
            },
            "courses": {
                "count": len(import_data.get("courses", [])),
                "samples": import_data.get("courses", [])[:3]
            },
            "prerequisites": {
                "count": len(import_data.get("prerequisites", [])),
                "samples": import_data.get("prerequisites", [])[:3]
            }
        }
        
        # Calculate total entities
        preview["total_entities"] = sum([
            preview[key]["count"] for key in preview if key != "total_entities"
        ])
        
        return preview
    
    async def execute_import(
        self, 
        import_data: Dict[str, Any], 
        user_id: int
    ) -> Dict[str, Any]:
        """
        Execute import as single transaction with audit logging.
        
        Requirements: 9.7, 9.8, 9.10
        
        Args:
            import_data: Dictionary with validated import data
            user_id: ID of user performing the import
            
        Returns:
            Dictionary with summary of created entities
            
        Raises:
            ValueError: If import fails, all changes are rolled back
        """
        try:
            # Start transaction (will be managed by SQLAlchemy session)
            created_entities = {
                "universities": [],
                "campuses": [],
                "programs": [],
                "university_programs": [],
                "tracks": [],
                "semesters": [],
                "teaching_units": [],
                "courses": [],
                "prerequisites": []
            }
            
            # Create name-to-entity mappings for referencing
            university_map = {}
            program_map = {}
            track_map = {}
            semester_map = {}
            teaching_unit_map = {}
            course_map = {}
            
            # 1. Create Universities
            for uni_data in import_data.get("universities", []):
                uni = University(
                    name=uni_data.get("name"),
                    name_de=uni_data.get("name_de"),
                    country=uni_data.get("country", "Germany"),
                    description=uni_data.get("description"),
                    description_de=uni_data.get("description_de")
                )
                self.db.add(uni)
                self.db.flush()  # Get ID without committing
                
                university_map[uni.name] = uni
                created_entities["universities"].append(uni.id)
                
                # Log creation
                await self.audit_service.log_create(
                    "university", uni.id, {"name": uni.name}, user_id
                )
            
            # 2. Create Campuses
            for campus_data in import_data.get("campuses", []):
                uni_name = campus_data.get("university_name")
                uni = university_map.get(uni_name)
                
                if not uni:
                    # Try to find in existing universities
                    uni = self.db.query(University).filter(
                        University.name == uni_name,
                        University.is_deleted == False
                    ).first()
                
                if uni:
                    campus = Campus(
                        university_id=uni.id,
                        name=campus_data.get("name"),
                        name_de=campus_data.get("name_de"),
                        location=campus_data.get("location"),
                        description=campus_data.get("description"),
                        description_de=campus_data.get("description_de")
                    )
                    self.db.add(campus)
                    self.db.flush()
                    
                    created_entities["campuses"].append(campus.id)
                    
                    await self.audit_service.log_create(
                        "campus", campus.id, {"name": campus.name}, user_id
                    )
            
            # 3. Create Programs
            for prog_data in import_data.get("programs", []):
                program = StudyProgram(
                    name=prog_data.get("name"),
                    name_de=prog_data.get("name_de"),
                    code=prog_data.get("code"),
                    description=prog_data.get("description"),
                    description_de=prog_data.get("description_de")
                )
                self.db.add(program)
                self.db.flush()
                
                program_map[program.name] = program
                created_entities["programs"].append(program.id)
                
                await self.audit_service.log_create(
                    "study_program", program.id, {"name": program.name}, user_id
                )
            
            # 4. Create University-Program links
            for link_data in import_data.get("university_programs", []):
                uni_name = link_data.get("university_name")
                prog_name = link_data.get("program_name")
                
                uni = university_map.get(uni_name) or self.db.query(University).filter(
                    University.name == uni_name, University.is_deleted == False
                ).first()
                
                program = program_map.get(prog_name) or self.db.query(StudyProgram).filter(
                    StudyProgram.name == prog_name, StudyProgram.is_deleted == False
                ).first()
                
                if uni and program:
                    stmt = university_programs.insert().values(
                        university_id=uni.id,
                        study_program_id=program.id,
                        created_at=datetime.now(timezone.utc)
                    )
                    self.db.execute(stmt)
                    created_entities["university_programs"].append(
                        f"{uni_name} <-> {prog_name}"
                    )
            
            # 5. Create Academic Tracks
            for track_data in import_data.get("tracks", []):
                prog_name = track_data.get("program_name")
                program = program_map.get(prog_name) or self.db.query(StudyProgram).filter(
                    StudyProgram.name == prog_name, StudyProgram.is_deleted == False
                ).first()
                
                if program:
                    level_str = track_data.get("level", "").lower()
                    level = TrackLevel.BACHELOR if level_str == "bachelor" else (
                        TrackLevel.MASTER if level_str == "master" else TrackLevel.DOCTORATE
                    )
                    
                    track = AcademicTrack(
                        study_program_id=program.id,
                        name=track_data.get("name"),
                        name_de=track_data.get("name_de"),
                        level=level,
                        total_ects_required=track_data.get("total_ects_required"),
                        description=track_data.get("description"),
                        description_de=track_data.get("description_de"),
                        graduation_conditions=track_data.get("graduation_conditions")
                    )
                    self.db.add(track)
                    self.db.flush()
                    
                    track_map[track.name] = track
                    created_entities["tracks"].append(track.id)
                    
                    await self.audit_service.log_create(
                        "academic_track", track.id, {"name": track.name}, user_id
                    )
            
            # 6. Create Semesters
            for sem_data in import_data.get("semesters", []):
                track_name = sem_data.get("track_name")
                track = track_map.get(track_name) or self.db.query(AcademicTrack).filter(
                    AcademicTrack.name == track_name, AcademicTrack.is_deleted == False
                ).first()
                
                if track:
                    semester = Semester(
                        academic_track_id=track.id,
                        name=sem_data.get("name"),
                        name_de=sem_data.get("name_de"),
                        semester_number=sem_data.get("semester_number"),
                        ects_required=sem_data.get("ects_required"),
                        description=sem_data.get("description"),
                        description_de=sem_data.get("description_de")
                    )
                    self.db.add(semester)
                    self.db.flush()
                    
                    semester_map[semester.name] = semester
                    created_entities["semesters"].append(semester.id)
                    
                    await self.audit_service.log_create(
                        "semester", semester.id, {"name": semester.name}, user_id
                    )
            
            # 7. Create Teaching Units
            for tu_data in import_data.get("teaching_units", []):
                sem_name = tu_data.get("semester_name")
                semester = semester_map.get(sem_name) or self.db.query(Semester).filter(
                    Semester.name == sem_name, Semester.is_deleted == False
                ).first()
                
                if semester:
                    teaching_unit = TeachingUnit(
                        semester_id=semester.id,
                        name=tu_data.get("name"),
                        name_de=tu_data.get("name_de"),
                        code=tu_data.get("code"),
                        ects_required=tu_data.get("ects_required"),
                        description=tu_data.get("description"),
                        description_de=tu_data.get("description_de")
                    )
                    self.db.add(teaching_unit)
                    self.db.flush()
                    
                    teaching_unit_map[teaching_unit.name] = teaching_unit
                    created_entities["teaching_units"].append(teaching_unit.id)
                    
                    await self.audit_service.log_create(
                        "teaching_unit", teaching_unit.id, {"name": teaching_unit.name}, user_id
                    )
            
            # 8. Create Courses
            for course_data in import_data.get("courses", []):
                sem_name = course_data.get("semester_name")
                semester = semester_map.get(sem_name) or self.db.query(Semester).filter(
                    Semester.name == sem_name, Semester.is_deleted == False
                ).first()
                
                if semester:
                    tu_name = course_data.get("teaching_unit_name")
                    teaching_unit = None
                    
                    if tu_name:
                        teaching_unit = teaching_unit_map.get(tu_name) or self.db.query(
                            TeachingUnit
                        ).filter(
                            TeachingUnit.name == tu_name, 
                            TeachingUnit.is_deleted == False
                        ).first()
                    
                    course = Course(
                        semester_id=semester.id,
                        teaching_unit_id=teaching_unit.id if teaching_unit else None,
                        name=course_data.get("name"),
                        name_de=course_data.get("name_de"),
                        code=course_data.get("code"),
                        ects_credits=course_data.get("ects_credits"),
                        coefficient=course_data.get("coefficient"),
                        difficulty_level=course_data.get("difficulty_level"),
                        description=course_data.get("description"),
                        description_de=course_data.get("description_de")
                    )
                    self.db.add(course)
                    self.db.flush()
                    
                    course_map[course.name] = course
                    created_entities["courses"].append(course.id)
                    
                    await self.audit_service.log_create(
                        "course", course.id, {"name": course.name}, user_id
                    )
            
            # 9. Create Prerequisites
            for prereq_data in import_data.get("prerequisites", []):
                course_name = prereq_data.get("course_name")
                prereq_name = prereq_data.get("prerequisite_name")
                
                course = course_map.get(course_name) or self.db.query(Course).filter(
                    Course.name == course_name, Course.is_deleted == False
                ).first()
                
                prerequisite = course_map.get(prereq_name) or self.db.query(Course).filter(
                    Course.name == prereq_name, Course.is_deleted == False
                ).first()
                
                if course and prerequisite:
                    stmt = course_prerequisites.insert().values(
                        course_id=course.id,
                        prerequisite_id=prerequisite.id,
                        created_at=datetime.now(timezone.utc)
                    )
                    self.db.execute(stmt)
                    created_entities["prerequisites"].append(
                        f"{course_name} <- {prereq_name}"
                    )
            
            # Commit transaction
            self.db.commit()
            
            # Create summary
            summary = {
                "success": True,
                "message": "Import completed successfully",
                "created_counts": {
                    key: len(value) for key, value in created_entities.items()
                },
                "total_created": sum(len(v) for v in created_entities.values()),
                "timestamp": datetime.now(timezone.utc).isoformat()
            }
            
            return summary
            
        except SQLAlchemyError as e:
            # Rollback on error
            self.db.rollback()
            raise ValueError(f"Import failed and was rolled back: {str(e)}")
        except Exception as e:
            # Rollback on any error
            self.db.rollback()
            raise ValueError(f"Import failed and was rolled back: {str(e)}")


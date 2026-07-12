"""
Unit tests for ImportService

Tests cover:
- Excel file parsing
- Validation (structural and semantic)
- Preview generation
- Transactional import execution
- Rollback on errors
- Integration with ValidationService and AuditService
"""
import pytest
from unittest.mock import Mock, patch, MagicMock, AsyncMock
from sqlalchemy.orm import Session
from datetime import datetime
import tempfile
import openpyxl

from app.services.import_service import ImportService, ImportError
from app.models.university import University
from app.models.study_program import StudyProgram
from app.models.academic_track import AcademicTrack, TrackLevel
from app.models.semester import Semester
from app.models.course import Course


@pytest.fixture
def mock_db():
    """Create mock database session"""
    db = Mock(spec=Session)
    db.query.return_value.filter.return_value.first.return_value = None
    db.query.return_value.filter.return_value.all.return_value = []
    db.query.return_value.count.return_value = 0
    return db


@pytest.fixture
def import_service(mock_db):
    """Create ImportService instance with mocked dependencies"""
    return ImportService(mock_db)


@pytest.fixture
def sample_import_data():
    """Sample import data for testing"""
    return {
        "universities": [
            {"name": "TU Berlin", "name_de": "Technische Universität Berlin", "country": "Germany", "_row": 2}
        ],
        "campuses": [
            {"university_name": "TU Berlin", "name": "Main Campus", "location": "Berlin", "_row": 2}
        ],
        "programs": [
            {"name": "Computer Science", "name_de": "Informatik", "code": "CS", "_row": 2}
        ],
        "university_programs": [
            {"university_name": "TU Berlin", "program_name": "Computer Science", "_row": 2}
        ],
        "tracks": [
            {"program_name": "Computer Science", "name": "Bachelor CS", "level": "bachelor", "total_ects_required": 180, "_row": 2}
        ],
        "semesters": [
            {"track_name": "Bachelor CS", "name": "S1", "semester_number": 1, "ects_required": 30, "_row": 2}
        ],
        "teaching_units": [
            {"semester_name": "S1", "name": "UE1", "code": "UE1", "ects_required": 15, "_row": 2}
        ],
        "courses": [
            {"semester_name": "S1", "teaching_unit_name": "UE1", "name": "Intro to Programming",
             "code": "CS101", "ects_credits": 6, "coefficient": 1.0, "difficulty_level": 2, "_row": 2}
        ],
        "prerequisites": []
    }


class TestImportServiceParsing:
    """Test Excel file parsing functionality"""
    
    def test_parse_excel_file_creates_import_data_structure(self, import_service):
        """Test that parse_excel_file creates correct data structure"""
        # Create temporary Excel file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = openpyxl.Workbook()
            
            # Create Universities sheet
            ws_uni = wb.active
            ws_uni.title = "Universities"
            ws_uni.append(["name", "name_de", "country", "description", "description_de"])
            ws_uni.append(["TU Berlin", "Technische Universität Berlin", "Germany", "Tech university", "Tech Uni"])
            
            # Create Programs sheet
            ws_prog = wb.create_sheet("Programs")
            ws_prog.append(["name", "name_de", "code", "description", "description_de"])
            ws_prog.append(["Computer Science", "Informatik", "CS", "CS Program", "CS Programm"])
            
            wb.save(tmp.name)
            
            # Test parsing
            result = import_service.parse_excel_file(tmp.name)
            
            assert "universities" in result
            assert "programs" in result
            assert len(result["universities"]) == 1
            assert result["universities"][0]["name"] == "TU Berlin"
            assert len(result["programs"]) == 1
            assert result["programs"][0]["code"] == "CS"
    
    def test_parse_excel_file_handles_missing_sheets(self, import_service):
        """Test that missing sheets result in empty lists"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Universities"
            ws.append(["name", "country"])
            ws.append(["TU Berlin", "Germany"])
            wb.save(tmp.name)
            
            result = import_service.parse_excel_file(tmp.name)
            
            assert result["programs"] == []
            assert result["courses"] == []
    
    def test_parse_excel_file_invalid_file_raises_error(self, import_service):
        """Test that invalid file path raises ValueError"""
        with pytest.raises(ValueError, match="Failed to open Excel file"):
            import_service.parse_excel_file("/nonexistent/file.xlsx")


class TestImportServiceValidation:
    """Test import data validation"""
    
    @pytest.mark.asyncio
    async def test_validate_universities_structure_requires_name(self, import_service):
        """Test that universities must have a name"""
        import_data = {
            "universities": [{"name": "", "country": "Germany", "_row": 2}],
            "programs": [], "tracks": [], "semesters": [], "courses": [],
            "campuses": [], "university_programs": [], "teaching_units": [], "prerequisites": []
        }
        
        is_valid, errors = await import_service.validate_import_data(import_data)
        
        assert not is_valid
        assert len(errors) > 0
        assert any("name is required" in e["message"].lower() for e in errors)
    
    @pytest.mark.asyncio
    async def test_validate_courses_structure_validates_ects(self, import_service):
        """Test that courses must have valid ECTS (1-30)"""
        import_data = {
            "universities": [], "programs": [], "tracks": [], "semesters": [],
            "campuses": [], "university_programs": [], "teaching_units": [], "prerequisites": [],
            "courses": [
                {"name": "Course 1", "semester_name": "S1", "ects_credits": 0,
                 "coefficient": 1.0, "difficulty_level": 2, "_row": 2},
                {"name": "Course 2", "semester_name": "S1", "ects_credits": 31,
                 "coefficient": 1.0, "difficulty_level": 2, "_row": 3}
            ]
        }
        
        is_valid, errors = await import_service.validate_import_data(import_data)
        
        assert not is_valid
        assert any("1 and 30" in e["message"] for e in errors)
    
    @pytest.mark.asyncio
    async def test_validate_courses_structure_validates_coefficient(self, import_service):
        """Test that courses must have valid coefficient (0.1-10.0)"""
        import_data = {
            "universities": [], "programs": [], "tracks": [], "semesters": [],
            "campuses": [], "university_programs": [], "teaching_units": [], "prerequisites": [],
            "courses": [
                {"name": "Course 1", "semester_name": "S1", "ects_credits": 6,
                 "coefficient": 0.05, "difficulty_level": 2, "_row": 2}
            ]
        }
        
        is_valid, errors = await import_service.validate_import_data(import_data)
        
        assert not is_valid
        assert any("0.1 and 10.0" in e["message"] for e in errors)
    
    @pytest.mark.asyncio
    async def test_validate_courses_structure_validates_difficulty(self, import_service):
        """Test that courses must have valid difficulty (1-5)"""
        import_data = {
            "universities": [], "programs": [], "tracks": [], "semesters": [],
            "campuses": [], "university_programs": [], "teaching_units": [], "prerequisites": [],
            "courses": [
                {"name": "Course 1", "semester_name": "S1", "ects_credits": 6,
                 "coefficient": 1.0, "difficulty_level": 6, "_row": 2}
            ]
        }
        
        is_valid, errors = await import_service.validate_import_data(import_data)
        
        assert not is_valid
        assert any("1 and 5" in e["message"] for e in errors)
    
    @pytest.mark.asyncio
    async def test_validate_track_structure_validates_level(self, import_service):
        """Test that tracks must have valid level"""
        import_data = {
            "universities": [], "programs": [], "semesters": [], "courses": [],
            "campuses": [], "university_programs": [], "teaching_units": [], "prerequisites": [],
            "tracks": [
                {"name": "Track 1", "program_name": "CS", "level": "invalid", "total_ects_required": 180, "_row": 2}
            ]
        }
        
        is_valid, errors = await import_service.validate_import_data(import_data)
        
        assert not is_valid
        assert any("bachelor, master, doctorate" in e["message"].lower() for e in errors)
    
    @pytest.mark.asyncio
    async def test_validate_prerequisite_chains_detects_cycles(self, import_service):
        """Test that circular prerequisite dependencies are detected"""
        import_data = {
            "universities": [], "programs": [], "tracks": [], "semesters": [],
            "campuses": [], "university_programs": [], "teaching_units": [],
            "courses": [
                {"name": "Course A", "semester_name": "S1", "ects_credits": 6,
                 "coefficient": 1.0, "difficulty_level": 2, "_row": 2},
                {"name": "Course B", "semester_name": "S1", "ects_credits": 6,
                 "coefficient": 1.0, "difficulty_level": 2, "_row": 3},
                {"name": "Course C", "semester_name": "S1", "ects_credits": 6,
                 "coefficient": 1.0, "difficulty_level": 2, "_row": 4}
            ],
            "prerequisites": [
                {"course_name": "Course A", "prerequisite_name": "Course B", "_row": 2},
                {"course_name": "Course B", "prerequisite_name": "Course C", "_row": 3},
                {"course_name": "Course C", "prerequisite_name": "Course A", "_row": 4}
            ]
        }
        
        is_valid, errors = await import_service.validate_import_data(import_data)
        
        assert not is_valid
        assert any("circular" in e["message"].lower() for e in errors)


class TestImportServicePreview:
    """Test preview generation functionality"""
    
    @pytest.mark.asyncio
    async def test_preview_import_returns_counts(self, import_service, sample_import_data):
        """Test that preview returns entity counts"""
        preview = await import_service.preview_import(sample_import_data)
        
        assert preview["universities"]["count"] == 1
        assert preview["programs"]["count"] == 1
        assert preview["tracks"]["count"] == 1
        assert preview["semesters"]["count"] == 1
        assert preview["courses"]["count"] == 1
        assert preview["total_entities"] > 0
    
    @pytest.mark.asyncio
    async def test_preview_import_includes_samples(self, import_service, sample_import_data):
        """Test that preview includes sample entities"""
        preview = await import_service.preview_import(sample_import_data)
        
        assert len(preview["universities"]["samples"]) > 0
        assert preview["universities"]["samples"][0]["name"] == "TU Berlin"


class TestImportServiceExecution:
    """Test import execution and transactional behavior"""
    
    @pytest.mark.asyncio
    async def test_execute_import_creates_entities(self, import_service, sample_import_data):
        """Test that execute_import creates entities in database"""
        # Mock database operations
        import_service.db.add = Mock()
        import_service.db.flush = Mock()
        import_service.db.commit = Mock()
        import_service.db.execute = Mock()
        
        # Mock audit service with async mock
        import_service.audit_service.log_create = AsyncMock(return_value=None)
        
        result = await import_service.execute_import(sample_import_data, user_id=1)
        
        assert result["success"] is True
        assert "created_counts" in result
        assert result["created_counts"]["universities"] > 0
        assert import_service.db.commit.called
    
    @pytest.mark.asyncio
    async def test_execute_import_rollback_on_error(self, import_service, sample_import_data):
        """Test that import rolls back on error"""
        # Mock database to raise error during commit
        import_service.db.add = Mock()
        import_service.db.flush = Mock()
        import_service.db.commit = Mock(side_effect=Exception("Database error"))
        import_service.db.rollback = Mock()
        import_service.audit_service.log_create = AsyncMock(return_value=None)
        
        with pytest.raises(ValueError, match="rolled back"):
            await import_service.execute_import(sample_import_data, user_id=1)
        
        assert import_service.db.rollback.called
    
    @pytest.mark.asyncio
    async def test_execute_import_logs_audit_entries(self, import_service, sample_import_data):
        """Test that import logs audit entries for created entities"""
        import_service.db.add = Mock()
        import_service.db.flush = Mock()
        import_service.db.commit = Mock()
        import_service.db.execute = Mock()
        import_service.audit_service.log_create = AsyncMock(return_value=None)
        
        await import_service.execute_import(sample_import_data, user_id=1)
        
        # Verify audit logging was called
        assert import_service.audit_service.log_create.called
        assert import_service.audit_service.log_create.call_count > 0


class TestImportErrorClass:
    """Test ImportError class"""
    
    def test_import_error_to_dict(self):
        """Test ImportError serialization"""
        error = ImportError(row=5, sheet="Courses", message="Invalid ECTS", error_type="validation")
        
        error_dict = error.to_dict()
        
        assert error_dict["row"] == 5
        assert error_dict["sheet"] == "Courses"
        assert error_dict["message"] == "Invalid ECTS"
        assert error_dict["type"] == "validation"


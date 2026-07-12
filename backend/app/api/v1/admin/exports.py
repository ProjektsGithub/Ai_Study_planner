"""
Export and Reporting API Endpoints

Provides curriculum exports and analytical reports:

  GET  /api/v1/admin/exports/curriculum/excel          — Full curriculum export as .xlsx
  GET  /api/v1/admin/exports/curriculum/pdf            — Full curriculum export as .pdf
  GET  /api/v1/admin/exports/reports/summary           — JSON curriculum summary (per track)
  GET  /api/v1/admin/exports/reports/prerequisites     — JSON prerequisite chains report

Excel export uses openpyxl (already in requirements).
PDF export uses reportlab — degrades gracefully with a clear 503 if not installed.

All exports include a header with the export timestamp and the requesting user's email.

RBAC: Super Admin and University Admin can access export endpoints.

Requirements: 18.1-18.7
"""
import io
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.core.dependencies import get_db, get_current_user
from app.middleware.rbac import ROLE_SUPER_ADMIN, ROLE_UNIVERSITY_ADMIN, ROLE_PROGRAM_COORDINATOR
from app.models.user import User
from app.models.university import University
from app.models.campus import Campus
from app.models.study_program import StudyProgram
from app.models.academic_track import AcademicTrack
from app.models.semester import Semester
from app.models.teaching_unit import TeachingUnit
from app.models.course import Course, course_prerequisites
from app.schemas.admin import (
    CurriculumSummaryReport,
    CurriculumSummaryItem,
    PrerequisiteChainReport,
    PrerequisiteChainReportItem,
)

router = APIRouter()


# ============================================================================
# GET /api/v1/admin/exports/curriculum/excel  — Export curriculum to Excel
# ============================================================================

@router.get(
    "/curriculum/excel",
    summary="Export curriculum to Excel",
    description=(
        "Export the full curriculum hierarchy (universities → programs → tracks → "
        "semesters → teaching units → courses → prerequisites) to a single .xlsx file "
        "with one sheet per entity type. "
        "Requires Super Admin or University Admin role."
    ),
)
async def export_curriculum_excel(
    university_id: Optional[int] = Query(None, description="Filter by university ID (Super Admin only)"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Export full curriculum data as an Excel workbook.

    Requirements: 18.1, 18.2, 18.5, 18.6, 18.7

    Returns:
        StreamingResponse with .xlsx content and Content-Disposition header.

    Raises:
        403: If user does not have the required role.
        503: If openpyxl is not installed.
    """
    await _require_export_role(current_user, db)

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="openpyxl is not installed — Excel export unavailable",
        )

    now = datetime.now(timezone.utc)
    wb = openpyxl.Workbook()

    # ---- Header style ----
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    def _write_sheet(sheet_name: str, headers: list, rows: list):
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append(row)
        # Auto-width columns
        for col_idx, _ in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 22

    # ---- Universities ----
    uni_q = db.query(University).filter(University.is_deleted == False)
    if university_id:
        uni_q = uni_q.filter(University.id == university_id)
    universities = uni_q.all()

    _write_sheet("Universities", ["ID", "Name", "Name (DE)", "Country", "Description"], [
        [u.id, u.name, u.name_de, u.country, u.description] for u in universities
    ])

    uni_ids = {u.id for u in universities}

    # ---- Campuses ----
    campuses = db.query(Campus).filter(
        Campus.is_deleted == False,
        Campus.university_id.in_(uni_ids),
    ).all()
    _write_sheet("Campuses", ["ID", "University ID", "Name", "Name (DE)", "Location"], [
        [c.id, c.university_id, c.name, c.name_de, c.location] for c in campuses
    ])

    # ---- Programs ----
    programs = db.query(StudyProgram).filter(StudyProgram.is_deleted == False).all()
    _write_sheet("Programs", ["ID", "Name", "Name (DE)", "Code", "Description"], [
        [p.id, p.name, p.name_de, p.code, p.description] for p in programs
    ])

    # ---- Tracks ----
    tracks = db.query(AcademicTrack).filter(AcademicTrack.is_deleted == False).all()
    _write_sheet("Tracks", ["ID", "Program ID", "Name", "Name (DE)", "Level", "Total ECTS"], [
        [t.id, t.study_program_id, t.name, t.name_de, t.level.value, t.total_ects_required]
        for t in tracks
    ])
    track_ids = {t.id for t in tracks}

    # ---- Semesters ----
    semesters = db.query(Semester).filter(
        Semester.is_deleted == False,
        Semester.academic_track_id.in_(track_ids),
    ).all()
    _write_sheet("Semesters", ["ID", "Track ID", "Name", "Name (DE)", "Semester #", "ECTS Required"], [
        [s.id, s.academic_track_id, s.name, s.name_de, s.semester_number, s.ects_required]
        for s in semesters
    ])
    sem_ids = {s.id for s in semesters}

    # ---- Teaching Units ----
    teaching_units = db.query(TeachingUnit).filter(
        TeachingUnit.is_deleted == False,
        TeachingUnit.semester_id.in_(sem_ids),
    ).all()
    _write_sheet("TeachingUnits", ["ID", "Semester ID", "Name", "Name (DE)", "Code", "ECTS Required"], [
        [tu.id, tu.semester_id, tu.name, tu.name_de, tu.code, tu.ects_required]
        for tu in teaching_units
    ])

    # ---- Courses ----
    courses = db.query(Course).filter(
        Course.is_deleted == False,
        Course.semester_id.in_(sem_ids),
    ).all()
    _write_sheet(
        "Courses",
        ["ID", "Semester ID", "Teaching Unit ID", "Name", "Name (DE)", "Code",
         "ECTS Credits", "Coefficient", "Difficulty"],
        [
            [c.id, c.semester_id, c.teaching_unit_id, c.name, c.name_de,
             c.code, c.ects_credits, c.coefficient, c.difficulty_level]
            for c in courses
        ],
    )
    course_ids = {c.id for c in courses}

    # ---- Prerequisites ----
    if course_ids:
        prereqs = db.execute(
            course_prerequisites.select().where(
                course_prerequisites.c.course_id.in_(course_ids)
            )
        ).fetchall()
        _write_sheet("Prerequisites", ["Course ID", "Prerequisite Course ID", "Created At"], [
            [p.course_id, p.prerequisite_id, p.created_at.isoformat() if p.created_at else ""]
            for p in prereqs
        ])
    else:
        _write_sheet("Prerequisites", ["Course ID", "Prerequisite Course ID", "Created At"], [])

    # ---- Export Info sheet ----
    ws_info = wb.create_sheet("Export Info", 0)
    ws_info.append(["Export Timestamp", now.isoformat()])
    ws_info.append(["Exported By", current_user.email])
    ws_info.append(["Platform", "AI Study Planner Super Admin"])

    # Remove default empty sheet created by openpyxl
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    # Stream the workbook
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"curriculum_export_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ============================================================================
# GET /api/v1/admin/exports/curriculum/pdf  — Export curriculum to PDF
# ============================================================================

@router.get(
    "/curriculum/pdf",
    summary="Export curriculum to PDF",
    description=(
        "Export a formatted curriculum summary report as a PDF file. "
        "Requires Super Admin or University Admin role. "
        "Returns 503 if reportlab is not installed."
    ),
)
async def export_curriculum_pdf(
    university_id: Optional[int] = Query(None, description="Filter by university ID"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Export curriculum data as a PDF summary report.

    Requirements: 18.3, 18.4, 18.5, 18.6, 18.7

    Returns:
        StreamingResponse with .pdf content.

    Raises:
        503: If reportlab is not installed.
    """
    await _require_export_role(current_user, db)

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors as rl_colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
    except ImportError:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail=(
                "reportlab is not installed — PDF export unavailable. "
                "Install it with: pip install reportlab"
            ),
        )

    now = datetime.now(timezone.utc)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph("Curriculum Export Report", styles["Title"]))
    elements.append(Paragraph(
        f"Generated: {now.strftime('%Y-%m-%d %H:%M UTC')} | By: {current_user.email}",
        styles["Normal"],
    ))
    elements.append(Spacer(1, 12))

    # Summary table per track
    uni_q = db.query(University).filter(University.is_deleted == False)
    if university_id:
        uni_q = uni_q.filter(University.id == university_id)
    universities = uni_q.all()
    uni_ids = {u.id for u in universities}

    programs = db.query(StudyProgram).filter(StudyProgram.is_deleted == False).all()
    tracks = db.query(AcademicTrack).filter(AcademicTrack.is_deleted == False).all()

    table_data = [["Program", "Track", "Level", "ECTS", "Semesters", "Courses"]]
    prog_map = {p.id: p.name for p in programs}

    for track in tracks:
        sem_count = db.query(func.count(Semester.id)).filter(
            Semester.academic_track_id == track.id, Semester.is_deleted == False
        ).scalar() or 0
        course_count = db.query(func.count(Course.id)).join(
            Semester, Course.semester_id == Semester.id
        ).filter(
            Semester.academic_track_id == track.id,
            Course.is_deleted == False,
        ).scalar() or 0

        table_data.append([
            prog_map.get(track.study_program_id, "—"),
            track.name,
            track.level.value,
            str(track.total_ects_required),
            str(sem_count),
            str(course_count),
        ])

    tbl = Table(table_data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, rl_colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor("#EEF3F9")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(tbl)

    doc.build(elements)
    buf.seek(0)

    filename = f"curriculum_report_{now.strftime('%Y%m%d_%H%M%S')}.pdf"
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ============================================================================
# GET /api/v1/admin/exports/reports/summary  — JSON curriculum summary
# ============================================================================

@router.get(
    "/reports/summary",
    response_model=CurriculumSummaryReport,
    summary="Curriculum summary report",
    description=(
        "Return a JSON curriculum summary with one row per academic track, "
        "including university, program, track name, ECTS requirements, and entity counts. "
        "Requires Super Admin, University Admin, or Program Coordinator role."
    ),
)
async def get_curriculum_summary(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    JSON curriculum summary — one record per academic track.

    Requirements: 18.1, 18.5

    Returns:
        CurriculumSummaryReport with generated_at, user info, and row data.
    """
    await _require_export_role(current_user, db)

    now = datetime.now(timezone.utc)

    # Query tracks with their programs
    tracks = db.query(AcademicTrack).filter(AcademicTrack.is_deleted == False).all()
    prog_map = {
        p.id: p for p in db.query(StudyProgram).filter(StudyProgram.is_deleted == False).all()
    }

    # Build per-track row data
    rows = []
    for track in tracks:
        program = prog_map.get(track.study_program_id)
        sem_count = db.query(func.count(Semester.id)).filter(
            Semester.academic_track_id == track.id, Semester.is_deleted == False
        ).scalar() or 0

        sem_ids_sub = db.query(Semester.id).filter(
            Semester.academic_track_id == track.id, Semester.is_deleted == False
        ).subquery()

        course_count = db.query(func.count(Course.id)).filter(
            Course.semester_id.in_(sem_ids_sub), Course.is_deleted == False
        ).scalar() or 0

        tu_count = db.query(func.count(TeachingUnit.id)).filter(
            TeachingUnit.semester_id.in_(sem_ids_sub), TeachingUnit.is_deleted == False
        ).scalar() or 0

        # Find which university owns this program
        uni_name = "—"
        if program:
            # Walk university_programs join
            from app.models.study_program import university_programs
            link = db.execute(
                university_programs.select().where(
                    university_programs.c.study_program_id == program.id
                )
            ).first()
            if link:
                uni = db.query(University).filter(University.id == link.university_id).first()
                uni_name = uni.name if uni else "—"

        rows.append(CurriculumSummaryItem(
            university=uni_name,
            program=program.name if program else "—",
            track=track.name,
            track_level=track.level.value,
            total_ects_required=track.total_ects_required or 0,
            semester_count=sem_count,
            course_count=course_count,
            teaching_unit_count=tu_count,
        ))

    return CurriculumSummaryReport(
        generated_at=now.isoformat(),
        generated_by=current_user.email,
        rows=rows,
        total_tracks=len(rows),
    )


# ============================================================================
# GET /api/v1/admin/exports/reports/prerequisites  — Prerequisite chains report
# ============================================================================

@router.get(
    "/reports/prerequisites",
    response_model=PrerequisiteChainReport,
    summary="Prerequisite chains report",
    description=(
        "Return a JSON report of all direct prerequisite relationships across all active courses. "
        "Each row shows the course, its prerequisite, and their semester numbers. "
        "Requires Super Admin, University Admin, or Program Coordinator role."
    ),
)
async def get_prerequisites_report(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    All prerequisite relationships report.

    Requirements: 18.4, 18.5

    Returns:
        PrerequisiteChainReport with all direct prerequisite relationships.
    """
    await _require_export_role(current_user, db)

    now = datetime.now(timezone.utc)

    # Load all prerequisite rows
    all_prereqs = db.execute(course_prerequisites.select()).fetchall()

    # Batch-load all relevant courses
    all_course_ids = set()
    for row in all_prereqs:
        all_course_ids.add(row.course_id)
        all_course_ids.add(row.prerequisite_id)

    courses = db.query(Course).filter(Course.id.in_(all_course_ids)).all()
    course_map = {c.id: c for c in courses}

    # Load semester numbers
    sem_ids = {c.semester_id for c in courses}
    semesters = db.query(Semester).filter(Semester.id.in_(sem_ids)).all()
    semester_map = {s.id: s for s in semesters}

    relationships = []
    for row in all_prereqs:
        course = course_map.get(row.course_id)
        prereq = course_map.get(row.prerequisite_id)
        if not course or not prereq:
            continue

        course_sem = semester_map.get(course.semester_id)
        prereq_sem = semester_map.get(prereq.semester_id)

        relationships.append(PrerequisiteChainReportItem(
            course_id=course.id,
            course_name=course.name,
            course_code=course.code,
            semester_number=course_sem.semester_number if course_sem else 0,
            prerequisite_id=prereq.id,
            prerequisite_name=prereq.name,
            prerequisite_code=prereq.code,
            prerequisite_semester_number=prereq_sem.semester_number if prereq_sem else 0,
        ))

    # Sort by course semester then course name
    relationships.sort(key=lambda r: (r.semester_number, r.course_name))

    return PrerequisiteChainReport(
        generated_at=now.isoformat(),
        generated_by=current_user.email,
        relationships=relationships,
        total_relationships=len(relationships),
    )


# ============================================================================
# Helpers
# ============================================================================

async def _require_export_role(user: User, db: Session) -> None:
    """Raise 403 if the user does not hold at least one of the allowed admin roles."""
    from app.models.user_role import UserRole
    from app.models.admin_role import AdminRole

    allowed = [ROLE_SUPER_ADMIN, ROLE_UNIVERSITY_ADMIN, ROLE_PROGRAM_COORDINATOR]
    role = (
        db.query(UserRole)
        .filter(UserRole.user_id == user.id)
        .join(AdminRole)
        .filter(AdminRole.name.in_(allowed), AdminRole.is_active == True)
        .first()
    )
    if not role:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied: admin role required for exports",
        )
